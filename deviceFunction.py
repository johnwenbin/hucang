import requests
import json
import os
from PIL import Image
import pandas
import time
import xlsxwriter
import openpyxl
from openpyxl.drawing.image import Image


class DeviceFunction:

    def __init__(self):
        # self.baseUrl = "http://140.249.154.2:9005/"
        self.baseUrl = 'http://192.168.1.64:8090/'
        # self.param = {'userID': 'admin', 'pwd': 'Safery001com!'}
        self.param = {'userID': 'admin', 'pwd': '312b248f2e48ff1e27dccc6a4552e3cb', 'dataSource': 'AZT_DEVICE'}         # 密码1111QQqq
        # self.param = {'userID': 'admin', 'pwd': '0db659ed1dc663d087a928853e07ec7e', 'dataSource': 'AZT_DEVICE'}

    # 获取token
    def getToken(self):
        login_path = "/v1/saferycom/n/login/encryptLogin"
        # token = json.loads(requests.post(self.baseUrl + login_path, self.param).text).get('result').get('token')
        json_context = json.loads(requests.post(self.baseUrl + login_path, self.param).text)
        code = json_context.get('code')
        if code == 500:
            time.sleep(5)
            token = json.loads(requests.post(self.baseUrl + login_path, self.param).text).get('result').get('token')
            print(token)
            return token
        else:
            token = json_context.get('result').get('token')
            print(token)
            return token


    # 获取截图文件夹目录
    def currentDir(self, filename):
        if os.path.exists(filename):
            filedir = os.path.abspath(filename)
            print(filedir)
            return filedir
        else:
            os.mkdir(filename)
            filedir = os.path.abspath(filename)
            print(filedir)
            return filedir

    # 获取截图图片名前缀
    def png_name(self, png_dir):
        png_name = []
        for root, dirs, files in os.walk(png_dir):
            for i in files:
                png_name.append(i.split('-')[0])
            print(png_name)
            return png_name

    # 获取ip值
    def getIp(self, ipname):
        row = 0
        I = []
        # wb = openpyxl.load_workbook('ipList.xlsx')
        # sheet = wb.active
        df = pandas.read_excel('ipList.xlsx')
        data = df.loc[:, ipname]
        print(data)
        # num = len(df)
        # while row < num:
        #     data = df.iloc[row, 0]
        #     row = row + 1
        #     print('data:' + data)
        #     if data == ipname:
        #         ina = df.iloc[[row], :]
        #         print(ina)
        #         return ina
        #         print('ina:' + ina)


    # 以下为excel插入图片测试代码
    def excel_image(self, png_name, filein, filedir):
        row = 0
        E = []
        wb = openpyxl.load_workbook(filein)
        sheet = wb.active
        num = sheet.max_row - 1
        print(num)
        new_size = (120, 80)
        df = pandas.read_excel(filein)
        while row < num:
            data = df.iloc[row, 0]
            row = row + 1
            E.append(data)
            print(data)
            if data in png_name:
                img = Image(filedir + "\\" + str(data) + '.png')
                img.width, img.height = new_size
                sheet.add_image(img, 'H' + str(row))
        wb.save(filein)
        # print(E)


if __name__ == '__main__':
    DeviceFunction().getToken()
    # DeviceFunction().currentDir('ScreenShot')
    # DeviceFunction().png_name('/DMAutoTestReport\ScreenShot')
    # DeviceFunction().getIp('福建生产')
    # DeviceFunction().excel_image([1, 2, 3, 4, 5, 6], "设备管理程序上线测试报告.xlsx", 'E:\deviceManage\DMAutoTestReport\ScreenShot')
